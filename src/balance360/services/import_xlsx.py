from datetime import datetime
from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.orm import Session

from balance360.crud import account as account_crud
from balance360.crud import import_batch as import_batch_crud
from balance360.crud import import_row as import_row_crud
from balance360.crud import transaction as transaction_crud
from balance360.enums import ImportRowStatus, TransactionType
from balance360.exceptions import ImportServiceError
from balance360.models.import_batch import ImportBatch
from balance360.schemas.import_batch import ImportBatchCreate
from balance360.schemas.import_row import ImportRowCreate
from balance360.schemas.transaction import TransactionCreate
from balance360.services.import_rule import extract_amount


def parse_sheet(ws: Worksheet) -> tuple[list[dict], list[dict]]:
    valid_rows = []
    review_rows = []

    for source_row, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        date_val, description, debit, credit, *rest = row

        if date_val == description == debit == credit is None:
            continue

        if description:
            description = str(description)
        else:
            description = "(sin descripcion)"

        valid = False
        reason = ""

        if not isinstance(date_val, datetime):
            reason = "Fecha invalida"

        elif not isinstance(debit, (int, float)) and not isinstance(credit, (int, float)):
            reason = "Importes invalidos"

        else:
            is_split_row = (isinstance(credit, (int, float)) and credit > 0) and (
                isinstance(debit, (int, float)) and debit > 0
            )

            if isinstance(credit, (int, float)) and credit > 0:
                valid_rows.append(
                    {
                        "date": date_val,
                        "description": description,
                        "amount": Decimal(str(credit)),
                        "transaction_type": TransactionType.income,
                        "source_row": source_row,
                    }
                )
                valid = True

            if isinstance(debit, (int, float)) and debit > 0:
                valid_rows.append(
                    {
                        "date": date_val,
                        "description": "Retencion impositiva :: " + description
                        if is_split_row
                        else description,
                        "amount": Decimal(str(debit)),
                        "transaction_type": TransactionType.expense,
                        "source_row": source_row,
                    }
                )
                valid = True

        if not valid:
            if isinstance(debit, (int, float)) or isinstance(credit, (int, float)):
                row_debit = Decimal(str(debit)) if isinstance(debit, (int, float)) else None
                row_credit = Decimal(str(credit)) if isinstance(credit, (int, float)) else None
            else:
                row_debit = extract_amount(description)  # solo acá, cuando no hay columna numérica
                row_credit = None

            review_rows.append(
                {
                    "date": date_val if isinstance(date_val, datetime) else None,
                    "description": description,
                    "debit": row_debit,
                    "credit": row_credit,
                    "source_row": source_row,
                    "reason": reason or "No hay credito ni debito",
                }
            )

    return (valid_rows, review_rows)


def import_workbook(
    db: Session,
    file_bytes: BytesIO,
    filename: str,
) -> ImportBatch:
    accounts = account_crud.get_all(db)

    if not accounts:
        raise ImportServiceError("No hay cuentas configuradas, crea una antes de importar")

    wb = load_workbook(file_bytes, data_only=True)

    accounts_to_process = [account for account in accounts if account.name in wb.sheetnames]

    if not accounts_to_process:
        raise ImportServiceError("Ninguna hoja del Excel coincide con tus cuentas")

    import_batch = import_batch_crud.create(db, ImportBatchCreate(filename=filename))

    for account in accounts_to_process:
        valid_rows, review_rows = parse_sheet(ws=wb[account.name])

        for valid_row in valid_rows:
            transaction_crud.create(
                db,
                TransactionCreate(
                    date=valid_row["date"],
                    description=valid_row["description"],
                    amount=valid_row["amount"],
                    type=valid_row["transaction_type"],
                    account_id=account.id,
                    source_file=filename,
                    source_sheet=account.name,
                    source_row=valid_row["source_row"],
                    import_batch_id=import_batch.id,
                ),
            )

        for review_row in review_rows:
            import_row_crud.create(
                db,
                ImportRowCreate(
                    batch_id=import_batch.id,
                    account_id=account.id,
                    source_row=review_row["source_row"],
                    date=review_row["date"],
                    description=review_row["description"],
                    debit=review_row["debit"],
                    credit=review_row["credit"],
                    status=ImportRowStatus.needs_review,
                    reason=review_row["reason"],
                ),
            )

    return import_batch
