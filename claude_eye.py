import openpyxl

archivos = [
    "Ventas 01.xlsx",
    "Compras  01.xlsx",
    "Planilla de Flujo 2026-01.xlsx",
    "Resultado 2026.xlsx",
]

for nombre in archivos:
    print(f"\n{'='*50}")
    print(f"ARCHIVO: {nombre}")
    wb = openpyxl.load_workbook(nombre, read_only=True, data_only=True)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        print(f"  Sheet: {sheet}")
        # primeras 3 filas para ver encabezados y muestra
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= 3:
                break
            print(f"    {row}")
    wb.close()