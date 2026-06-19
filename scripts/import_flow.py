import uuid
from decimal import Decimal
from openpyxl import load_workbook, Workbook
from datetime import datetime
from sqlalchemy import select
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.orm import Session
from balance360.models import Account, ImportRule, Transaction
from balance360.enums import TransactionType
from balance360.database import SessionLocal
from balance360.services.import_rule import find_best_rule, extract_amount

def clean_work_book(wb: Workbook, ws_name: str):
    if ws_name in wb.sheetnames: del wb[ws_name]
    ws = wb.create_sheet(ws_name)
    headers = ['Fecha', 'Descripcion', 'Debito', 'Credito', 'Cuenta']
    ws.append(headers)

def export_rows(wb: Workbook, rows, sheet: str):
    ws = wb[sheet]    
    for row in rows:
        ws.append([row['date'], row['description'], row['debit'], row['credit'], row['account']])
    
def parse_sheet(wb: Workbook, ws_name: str) -> tuple[list[dict], list[dict], list[dict], list[dict]]:
    ws = wb[ws_name]
    valid_rows = []
    skipped_rows = []
    cash_rows = []
    visa_rows = []
    
    for source_row, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        date_val, description, debit, credit, *rest = row
        if description:
            description = str(description)
        else:
            description = '(sin descripcion)'

        if not isinstance(date_val, datetime):
            skipped_rows.append({
                'date': date_val, 
                'description': description,
                'debit': debit,
                'credit': credit,
                'source_row': source_row
            })
        elif not isinstance(debit, (int, float)) and not isinstance(credit, (int, float)):
            if 'VISA' in description:
                amount = extract_amount(description)
                if amount is not None:
                    visa_rows.append({
                        'date': date_val, 
                        'description': description,
                        'debit': amount,
                        'credit': Decimal(0),
                        'source_row': source_row,
                        'account': ws_name,
                        'amount': amount,
                        'transaction_type': TransactionType.expense
                    })
                else:
                    skipped_rows.append({
                        'date': date_val, 
                        'description': description,
                        'debit': Decimal(0),
                        'credit': Decimal(0),
                        'source_row': source_row,
                        'account': ws_name
                    })
            else:
                amount = extract_amount(description)
                cash_rows.append({
                    'date': date_val, 
                    'description': description,
                    'debit': amount if amount is not None else Decimal(0),
                    'credit': Decimal(0),
                    'source_row': source_row,
                    'account': ws_name,
                })
        else:
            is_split_row = (isinstance(credit, (int, float)) and credit > 0) and (isinstance(debit, (int, float)) and debit > 0)

            matched = False
            if isinstance(credit, (int, float)) and credit > 0:
                valid_rows.append({
                    'date': date_val,
                    'description': description,
                    'amount': Decimal(str(credit)),
                    'transaction_type': TransactionType.income,
                    'source_row': source_row
                })
                matched = True
            if isinstance(debit, (int, float)) and debit > 0:
                valid_rows.append({
                    'date': date_val,
                    'description': 'Retencion impositiva :: '+description if is_split_row else description,
                    'amount': Decimal(str(debit)),
                    'transaction_type': TransactionType.expense,
                    'source_row': source_row
                })
                matched = True
            if not matched:
                skipped_rows.append({
                    'date': date_val, 
                    'description': description,
                    'debit': debit,
                    'credit': credit,
                    'source_row': source_row
                })

    return (valid_rows, skipped_rows, cash_rows, visa_rows)

def load_rules(db: Session) -> list[ImportRule]:
    rules = db.execute(select(ImportRule)).scalars().all()
    return list(rules)

def load_accounts(db: Session) -> list[Account]:
    accounts = db.execute(select(Account)).scalars().all()
    return list(accounts)

def import_sheet(db: Session, rows: list[dict], account: Account, rules: list[ImportRule], source_file: str):
    for row in rows:
        amount = row['amount']
        
        import_rule = find_best_rule(row['description'], row['transaction_type'], rules)
        
        transaction_dict = {
            'id': uuid.uuid4(),
            'date': row['date'],
            'description': row['description'],
            'amount': amount,
            'type': row['transaction_type'],
            'account_id': account.id,
            'entity_id': import_rule.entity_id if import_rule else None,
            'contact_id': import_rule.contact_id if import_rule else None,
            'category_id': import_rule.category_id if import_rule else None,
            'is_manual': False,
            'is_transfer': import_rule.is_transfer if import_rule else False,
            'applied_rule_id': import_rule.id if import_rule else None,
            'source_file': source_file,
            'source_row': row['source_row']
        }

        stmt = pg_insert(Transaction).values(**transaction_dict).on_conflict_do_nothing(
            constraint="uq_transaction"
        )
        db.execute(stmt)
    db.commit()

if __name__ == "__main__":
    wb_name = 'Planilla de Flujo 2026-01.xlsx'
    wb = load_workbook(wb_name, data_only=True)

    with SessionLocal() as db:
        rules = load_rules(db)
        accounts = load_accounts(db)
        visa_account = next(a for a in accounts if a.name == 'VISA Ciudad')
        accounts_to_import = ['Banco Frances (ARS)', 'Banco Ciudad (ARS)', 'MP']
        
        clean_work_book(wb, 'Efectivo')
        clean_work_book(wb, 'VISA Ciudad')

        for account in accounts:
            if account.name not in accounts_to_import:
                continue
            source_file = wb_name + "::" + account.name
            valid_rows, skipped_rows, cash_rows, visa_rows = parse_sheet(wb, account.name)
            if cash_rows: export_rows(wb, cash_rows, 'Efectivo')
            if visa_rows: export_rows(wb, visa_rows, 'VISA Ciudad')

            print(f"Solapa {account.name}: {len(valid_rows)} filas validas, {len(skipped_rows)} filas saltadas, {len(cash_rows)} filas de efectivo creadas, {len(visa_rows)} filas VISA creadas")
            import_sheet(db, valid_rows, account, rules, source_file)
            if visa_rows:
                import_sheet(db, visa_rows, visa_account, rules, source_file)
        wb.save(wb_name+"- procesado")

        
